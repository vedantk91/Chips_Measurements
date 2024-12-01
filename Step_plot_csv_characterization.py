from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt

file_path = 'Recording2.xlsx'

def read_data_from_csv(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    print(ws)
    data = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0] is not None and row[0] < 256:  
            data.append(row[0])
        if len(data) == 500:  
            break
    return data

data = read_data_from_csv(file_path)

def generate_step_diagram(data):
    x = [0]  
    y = [data[0]]  
    for i in range(1, len(data)):
        x.append(x[-1])  
        y.append(data[i])
        x.append(x[-1] + 1)  
        y.append(data[i])
    return x, y

x, y = generate_step_diagram(data)

plt.plot(x, y, drawstyle='steps-pre', color='black')
plt.show()